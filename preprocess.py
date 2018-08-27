# -*- coding: utf-8 -*-
import jieba
import openpyxl
from collections import Counter
import gensim
from numpy.random import rand
import re

#stop words
stpwrdpath = "data/stop_words.txt"
stpwrd_dic = open(stpwrdpath, 'rb')
stpwrd_content = stpwrd_dic.read()
stpwrdlst = stpwrd_content.splitlines()
stpwrdlst = [wrd.decode('GBK', 'ignore') for wrd in stpwrdlst]
stpwrd_dic.close()

#emoji
emoji = re.compile("["
        u"\U0001F600-\U0001F64F"  # emoticons
        u"\U0001F300-\U0001F5FF"  # symbols & pictographs
        u"\U0001F680-\U0001F6FF"  # transport & map symbols
        u"\U0001F1E0-\U0001F1FF"  # flags (iOS)
        u"U+2605-U+2b52"                 
        u"U+66d"           "]+", flags=re.UNICODE)
#digits & english & punctuation
dig_eng = re.compile("[\d\w]*")
punc1 = [".","_","(",")","/","-","~",'"',"'","#","...","!","....","*","&"]
punc2 = [u"【",u"】",u"（",u"）",u"…",u"—",u"·",u"～",
         u"#",u"／",u"★",u"❤️️️️️️️",u"」",u"「",u"→",u"＂",u"“",u"ಥ",u"❤"]
punc3 = stpwrdlst[:13]
#vocabulary
model = gensim.models.KeyedVectors.load_word2vec_format('data/cn.skipgram.bin',
                                                        binary=True, unicode_errors='ignore')
vocab = model.vocab

mydic = "data/dict.txt"
jieba.load_userdict(mydic)

label_path = "data/douban-sentiment-labeled.xlsx"
book = openpyxl.load_workbook(label_path)
sheet = book.active
def convert(polarity):
    if polarity=="positive":
        return 0
    elif polarity=="negtive":
        return 2
    #elif polarity=="none":
    #    return 3
    else:
        return 1

sentence_all = set()
word_all = []
data_tr = []
data_te = []
for i in range(2,1354):
    name = sheet.cell(row=i, column=2).value
    plot = sheet.cell(row=i, column=4).value
    cast = sheet.cell(row=i, column=5).value
    review = sheet.cell(row=i, column=7).value
    #document_decode = review.decode('GBK', 'ignore')
    document_decode = review
    if document_decode is None:
        continue
    document_cut = jieba.cut(document_decode)
    # segmentation into sentences
    #document_cut = list(document_cut)
    # remove stop words
    #document_cut = [wrd for wrd in document_cut
    #                if wrd not in stpwrdlst and wrd!=' ']
    #remove emojis
    document_cut = ' '.join(document_cut)
    document_cut = emoji.sub(r'', document_cut)
    #remove digits
    document_cut = dig_eng.sub(r'', document_cut)
    #remove double spaces
    document_cut = document_cut.replace('  ',' ')
    #remove duplicate revies
    if document_cut not in sentence_all:
        sentence_all.add(document_cut)
    else:
        continue
    document_cut = document_cut.split(' ')
    #remove '' and ' '
    document_cut = [wrd for wrd in document_cut if wrd!='' and wrd!=' ']
    # remove punctuations
    document_cut = [wrd for wrd in document_cut if wrd not in punc1
                    and wrd not in punc2 and wrd not in punc3]
    # remove too short reviews
    if len(document_cut) < 4:
        continue
    word_all += document_cut
    result = ' '.join(document_cut)
    result = result.encode('utf-8')
    r = rand()
    if(r>0.8):
        data_te.append([result,'plot',convert(plot)])
        data_te.append([result, 'cast', convert(cast)])

    data_tr.append([result, 'plot', convert(plot)])
    data_tr.append([result, 'cast', convert(cast)])

word_count = Counter(word_all)
print len(word_count)
count = word_count.most_common(3000)
cnt = [w for w,_ in count if w in vocab]
print len(cnt)
d=dict()#avoid including the same word twice
id=0
with open("data/word_id", 'w+') as f:
    for w,_ in count:
        if w==' ':
            continue
        if d.has_key(w.lower()):
            continue
        d[w.lower()]=id
        f.write(w.encode('utf-8') + ' ')
        f.write(str(id) + '\n')
        id+=1
        f.flush()

test_path = "data/douban_test"
train_path = "data/douban_train"
with open(test_path, 'w+') as te:
    for dat in data_te:
        te.write(dat[0] + '\n')
        te.write(dat[1] + '\n')
        te.write(str(dat[2]) + '\n')
        te.flush()
with open(train_path, 'w+') as tr:
    for dat in data_tr:
        tr.write(dat[0] + '\n')
        tr.write(dat[1] + '\n')
        tr.write(str(dat[2]) + '\n')
        tr.flush()

